import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.models import SearchSession

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
ALT_FILL = PatternFill("solid", fgColor="EBF3FB")
YELLOW_FILL = PatternFill("solid", fgColor="FFF9C4")
TOTAL_FILL = PatternFill("solid", fgColor="2E74B5")
WHITE_FONT = Font(color="FFFFFF", bold=True)
TOTAL_FONT = Font(color="FFFFFF", bold=True)

PATENT_COLS = [
    "Query ID", "Descrição", "Fonte", "Número", "Título", "Depositante",
    "Data Pub.", "Jurisdição", "IPC", "Resumo", "Lens ID",
    "Relevante? S/N", "Motivo Exclusão",
]

SUMARIO_COLS = [
    "Query ID", "Descrição", "Fonte", "String de Busca",
    "Total (base)", "Coletados", "Excluídos (triagem)", "Selecionados",
]


def _set_header_row(ws, cols: list[str]):
    for col_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 60)


def build_excel(session: SearchSession) -> bytes:
    wb = Workbook()

    # ── Aba 1: Patentes Encontradas ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Patentes Encontradas"
    _set_header_row(ws1, PATENT_COLS)

    for row_idx, patent in enumerate(session.results, start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
        values = [
            patent.query_id, patent.descricao, patent.fonte, patent.numero,
            patent.titulo, patent.depositante, patent.data_pub, patent.jurisdicao,
            patent.ipc, patent.resumo, patent.lens_id,
            patent.relevante, patent.motivo_exclusao,
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            if col_idx in (12, 13):
                cell.fill = YELLOW_FILL
            elif row_idx % 2 == 0:
                cell.fill = ALT_FILL

    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = ws1.dimensions
    _auto_width(ws1)

    # ── Aba 2: Sumário PRISMA ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Sumário PRISMA")
    _set_header_row(ws2, SUMARIO_COLS)

    total_coletados = 0
    for row_idx, item in enumerate(session.sumario, start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
        values = [
            item.query_id, item.descricao, item.fonte, item.string,
            item.total, item.coletado, item.excluidos, item.selecionados,
        ]
        total_coletados += item.coletado
        for col_idx, val in enumerate(values, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            if col_idx in (7, 8):
                cell.fill = YELLOW_FILL
            elif row_idx % 2 == 0:
                cell.fill = ALT_FILL

    total_row = len(session.sumario) + 2
    total_values = ["TOTAL", "", "", "", "", total_coletados, "[preencher]", "[preencher]"]
    for col_idx, val in enumerate(total_values, start=1):
        cell = ws2.cell(row=total_row, column=col_idx, value=val)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT

    _auto_width(ws2)

    # ── Aba 3: Critérios de Triagem ──────────────────────────────────────────
    ws3 = wb.create_sheet("Critérios de Triagem")
    criterios = """MARCAR S — RELEVANTE:
  A patente descreve sistema de informação, software ou método computacional
  aplicado à gestão, rastreabilidade ou suporte à decisão em CME/CSSD/SPD

MARCAR N — EXCLUÍDA (indicar código):
  EX1 — Equipamento físico de esterilização sem componente de software
  EX2 — Insumo ou produto consumível (indicadores, embalagens)
  EX3 — Método químico/físico sem sistema de informação
  EX4 — Fora do contexto hospitalar/médico-cirúrgico
  EX5 — Duplicata de outra query ou fonte
  EX6 — Idioma fora de PT/EN/ES

FONTES CONSULTADAS:
  Lens.org  — https://lens.org (token acadêmico gratuito)
  EPO OPS   — https://developers.epo.org (gratuito até 4GB/mês)
  INPI pePI — https://busca.inpi.gov.br/pePI (acesso anônimo público)

REFERÊNCIA METODOLÓGICA:
  MOHER, D. et al. PRISMA Statement. PLoS Medicine, v.6, n.7, e1000097, 2009."""

    for line_idx, line in enumerate(criterios.split("\n"), start=1):
        ws3.cell(row=line_idx, column=1, value=line)

    ws3.column_dimensions["A"].width = 80

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def build_csv(session: SearchSession) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(PATENT_COLS)
    for patent in session.results:
        writer.writerow([
            patent.query_id, patent.descricao, patent.fonte, patent.numero,
            patent.titulo, patent.depositante, patent.data_pub, patent.jurisdicao,
            patent.ipc, patent.resumo, patent.lens_id,
            patent.relevante, patent.motivo_exclusao,
        ])
    return ("﻿" + buf.getvalue()).encode("utf-8")
